from openpyxl import Workbook, load_workbook
import tweepy
from time import sleep

consumer_key = '8ZgghcwUoDm6gY8YYtlDSrzjH'
consumer_secret = 'RLSdYYGduGokDJ4nsKmhmumEpF99NNoqaQOrcLsxdCBWgYKk3k'
access_token_key = '1156170455147057152-6AC3AjpyEM1ixgRRulp04HrU04hU8L'
access_token_secret = 'jZLCfV7uQKeRU5CcHOgiFba26vF3Mm2Pxtx1RoznHa6rC'

auth = tweepy.OAuthHandler(consumer_key, consumer_secret)
auth.set_access_token(access_token_key, access_token_secret)

api = tweepy.API(auth)

try:
    workbook = load_workbook('data/data.xlsx')
    sheet = workbook.active
except:
    workbook = Workbook()
    sheet = workbook.active
    
    sheet["A1"] = "Phrases"
    sheet["B1"] = "Depressed"


def TweetCollection(thread, label):
    global index
    results = api.search(q=f"#{thread} -filter:retweets", count=1000, tweet_mode="extended", include_rts = False, lang="en")
    
    for tweet in results:
        sheet[f"A{index}"] = tweet.full_text
        sheet[f"B{index}"] = label
        index = index + 1
        
threads = [("depressed",1), ("happy",0), ("sad",1), ("joyful",0)]

index = sheet.max_row + 1
print(sheet.max_row + 1)
for thread, label in threads:
    TweetCollection(thread, label)
    
print(sheet.max_row)
    
while True:
    try:
        workbook.save(filename="data/data.xlsx")
        break
    except:
        print("Currently busy, trying again...")
        sleep(5)